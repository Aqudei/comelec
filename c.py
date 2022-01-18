import json
from typing import Optional, OrderedDict
from openpyxl.worksheet.worksheet import Worksheet
import requests
from seleniumrequests import Chrome
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
import csv
import json
import re
import shutil
import openpyxl
import argparse
import itertools
import pandas as pd
import numpy as np

PL = [
    'BAYAN MUNA',
    'ACT TEACHERS',
    'GABRIELA',
    'KABATAAN',
    "ANAKPAWIS"
]


def get_provinces(driver: Chrome):
    """
    docstring
    """
    response = driver.request(
        "GET", "https://comelec.gov.ph/2019NLEResults/data/regions/0/8.json")
    print(f"Status Code:{response.status_code}")
    pass


def clean(pl_name):
    """
    docstring
    """
    # rgx = re.compile(r"(\d+\s+)(.+)")
    # match = rgx.search(pl_name)
    # return match.group(2)
    return pl_name


def agg_by_province(input):
    wb = openpyxl.load_workbook(input)
    ws = wb.worksheets[0]
    ws1 = wb.create_sheet("By Province")
    rows = sorted(ws.iter_rows(min_row=2, values_only=True),
                  key=lambda x: (x[0], x[4]))
    grows = itertools.groupby(rows, key=lambda x: (x[0], x[4]))
    ws1.append(['Province', 'Partylist', 'Votes', 'Voters'])
    for k, g in grows:
        items = list(g)
        total_votes = sum([item[5] for item in items])
        total_voters = sum([item[6] for item in items])
        ws1.append([k[0], k[1], total_votes, total_voters])

    wb.save(input)
    wb.close()


def pivot_2(input):
    """
    docstring
    """
    pass


def agg_by_citymun(input):
    wb = openpyxl.load_workbook(input)
    ws = wb.worksheets[0]
    ws1 = wb.create_sheet("By City or Municipality")
    rows = sorted(ws.iter_rows(min_row=2, values_only=True),
                  key=lambda x: (x[0], x[1], x[4]))
    grows = itertools.groupby(rows, key=lambda x: (x[0], x[1], x[4]))
    ws1.append(['Province', 'City/Municipalities',
               'Partylist', 'Votes', 'Voters'])
    for k, g in grows:
        items = list(g)
        total_votes = sum([item[5] for item in items])
        total_voters = sum([item[6] for item in items])
        ws1.append([k[0], k[1], k[2], total_votes, total_voters])

    wb.save(input)
    wb.close()


def agg_by_brgys(input):
    wb = openpyxl.load_workbook(input)
    ws = wb.worksheets[0]
    ws1 = wb.create_sheet("By Barangays")
    rows = sorted(ws.iter_rows(min_row=2, values_only=True),
                  key=lambda x: (x[0], x[1], x[2], x[4]))
    grows = itertools.groupby(rows, key=lambda x: (x[0], x[1], x[2], x[4]))
    ws1.append(['Province', 'City/Municipalities',
               'Barangay', 'Partylist', 'Votes', 'Voters'])
    for k, g in grows:
        items = list(g)
        total_votes = sum([item[5] for item in items])
        total_voters = sum([item[6] for item in items])
        ws1.append([k[0], k[1], k[2], k[3], total_votes, total_voters])

    wb.save(input)
    wb.close()


def main(input_folder):

    with open("./output.csv", "wt", newline='') as fo:
        writer = csv.DictWriter(fo, fieldnames=(
            'dprovince', 'dmunicipality', 'dbarangay', 'dprecint', 'dpartyName', 'dvote', 'dtotal', 'dpercentage'))
        writer.writeheader()

        for root, dirs, files in os.walk(input_folder):
            for f in files:
                if not f.endswith(".json"):
                    continue

                with open(os.path.join(root, f), 'rt') as fp:
                    js = json.loads(fp.read())

                    for item in js:
                        item['dpartyName'] = clean(item['dpartyName'])
                        writer.writerow(item)

        fo.flush()


def rename():
    """
    docstring
    """
    for root, dirs, files in os.walk('./data'):
        for f in files:
            if f.endswith('.js'):
                fn, ext = os.path.splitext(f)
                dest = os.path.join(root, fn + '.json')
                shutil.move(os.path.join(root, f), dest)


def namefix(input, sheet, col):
    """
    docstring
    """
    rgx = re.compile(r'(\d+\s+)(.+)')
    wb = openpyxl.load_workbook(input)
    ws = wb.worksheets[sheet]
    for row in ws.iter_rows(min_row=2):
        match = rgx.search(row[col].value)
        newval = row[col].value
        if match:
            newval = match.group(2)

        row[col].value = newval

    wb.save(input)
    wb.close()


def pivot(input):
    """
    docstring
    """
    H = ['Province', 'Municipality', 'Barangay', 'Precint', 'Total Voters']
    with open('pivot.csv', 'wt', newline='') as fo:
        writer = csv.writer(fo)
        with open(input, 'rt', newline='') as fp:
            reader = csv.reader(fp)
            next(reader)
            items = itertools.groupby(
                [r for r in reader], key=lambda k: (k[0], k[1], k[2], k[3], k[6]))

            no_header = True
            for idx, (g, v) in enumerate(items):
                print(f"Processing line {idx+1}...")
                templ = data_template()
                if no_header:
                    header = [h for h in itertools.chain(
                        H, templ.keys())]
                    writer.writerow(header)
                    no_header = False

                for party in v:
                    d = {party[4]: party[5]}

                    for k in d.keys():
                        templ[k] = d[k]
                # data = [d[key] for key in sorted(d.keys())]
                writer.writerow(
                    [r for r in itertools.chain(g, templ.values())])


template = None


def data_template():
    """
    docstring
    """
    global template

    if not template:
        with open('./5567.json', 'rt') as jf:
            template = OrderedDict({
                clean(bo['bon']): 0 for bo in json.loads(jf.read())['bos']})

    return template


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--rename', action='store_true')
    parser.add_argument('--input')
    parser.add_argument('--pivot', action='store_true')
    parser.add_argument('--agg0', action='store_true')
    parser.add_argument('--agg1', action='store_true')
    parser.add_argument('--agg2', action='store_true')
    parser.add_argument('--agg3', action='store_true')
    parser.add_argument('--namefix', action='store_true')
    parser.add_argument('--col', type=int)
    parser.add_argument('--sheet', type=int)

    options = parser.parse_args()

    if options.rename:
        rename()

    if options.agg0:
        main(options.input)
        
    if options.agg1:
        agg_by_brgys(options.input)

    if options.pivot:
        pivot(options.input)



    # if options.agg2:
    #     agg_by_citymun(options.input)

    # if options.agg3:
    #     agg_by_province(options.input)

    # if options.namefix:
    #     namefix(options.input, options.sheet, options.col)
